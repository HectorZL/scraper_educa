from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
import difflib
import re
import unicodedata

import openpyxl

LANGUAGE_SUBJECT = "LENGUA Y LITERATURA"
FUZZY_MATCH_THRESHOLD = 0.85
FUZZY_MATCH_MARGIN = 0.15


class GradeFileError(ValueError):
    """Indica que un archivo de notas no tiene la estructura esperada."""


class NameMatchError(ValueError):
    """Indica que un nombre no puede relacionarse de manera única y segura."""


@dataclass(frozen=True)
class GradeRecord:
    row_number: int
    name: str
    normalized_name: str
    notes: tuple[str | None, str | None, str | None, str | None]


@dataclass(frozen=True)
class MatchResult:
    record: GradeRecord
    method: str
    score: float


def normalize_name(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").lower())
    text = "".join(
        character
        for character in text
        if not unicodedata.combining(character)
        and (character.isalnum() or character.isspace())
    )
    return re.sub(r"\s+", " ", text).strip()


def normalize_subject(value: object) -> str:
    return normalize_name(value)


def format_grade(
    value: object,
    row_number: int,
    column_name: str,
    allow_empty: bool,
) -> str | None:
    if value is None or str(value).strip() == "":
        if allow_empty:
            return None
        raise GradeFileError(
            f"Falta {column_name} en la fila {row_number}. Se requieren cuatro notas."
        )

    try:
        decimal_value = Decimal(str(value).strip().replace(",", "."))
    except InvalidOperation as error:
        raise GradeFileError(
            f"{column_name} no es numérica en la fila {row_number}."
        ) from error

    if not 0 <= decimal_value <= 10:
        raise GradeFileError(
            f"{column_name} está fuera del rango 0-10 en la fila {row_number}."
        )

    formatted = format(decimal_value.normalize(), "f")
    return formatted.rstrip("0").rstrip(".") if "." in formatted else formatted


def read_grade_records(
    file_path: str | Path,
    *,
    allow_partial_notes: bool = False,
) -> list[GradeRecord]:
    path = Path(file_path)
    if not path.is_file():
        raise GradeFileError(f"No existe el archivo de notas: {path}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active

    headers = {
        normalize_name(cell.value): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }
    required_headers = ["nombre", "nota 1", "nota 2", "nota 3", "nota 4"]
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        raise GradeFileError(
            f"Faltan columnas en {path.name}: {', '.join(missing_headers)}"
        )

    records = []
    seen_names = set()
    for row_number in range(2, sheet.max_row + 1):
        name_value = sheet.cell(row_number, headers["nombre"]).value
        normalized_name = normalize_name(name_value)
        if not normalized_name:
            continue
        if normalized_name in seen_names:
            raise GradeFileError(
                f"Nombre duplicado en {path.name}, fila {row_number}."
            )

        notes = tuple(
            format_grade(
                sheet.cell(row_number, headers[f"nota {index}"]).value,
                row_number,
                f"Nota {index}",
                allow_partial_notes,
            )
            for index in range(1, 5)
        )
        records.append(
            GradeRecord(
                row_number=row_number,
                name=str(name_value).strip(),
                normalized_name=normalized_name,
                notes=notes,
            )
        )
        seen_names.add(normalized_name)

    workbook.close()
    if not records:
        raise GradeFileError(f"No se encontraron estudiantes en {path.name}.")
    return records


def token_similarity(query: str, candidate: str) -> float:
    query_tokens = query.split()
    candidate_tokens = candidate.split()
    if not query_tokens or not candidate_tokens:
        return 0.0

    return sum(
        max(
            difflib.SequenceMatcher(None, query_token, candidate_token).ratio()
            for candidate_token in candidate_tokens
        )
        for query_token in query_tokens
    ) / len(query_tokens)


def find_unique_match(query_name: object, records: list[GradeRecord]) -> MatchResult:
    query = normalize_name(query_name)
    # El formulario Aportes antepone el número de orden al nombre del estudiante.
    query = re.sub(r"^\d+\s+", "", query).strip()
    if not query:
        raise NameMatchError("El nombre a comparar está vacío.")

    exact_matches = [record for record in records if record.normalized_name == query]
    if len(exact_matches) == 1:
        return MatchResult(exact_matches[0], "exact", 1.0)

    query_tokens = set(query.split())
    subset_matches = [
        record
        for record in records
        if query_tokens and query_tokens.issubset(set(record.normalized_name.split()))
    ]
    if len(subset_matches) == 1:
        return MatchResult(subset_matches[0], "token_subset", 1.0)
    if len(subset_matches) > 1:
        raise NameMatchError("El nombre abreviado coincide con más de un estudiante.")

    ranked = sorted(
        (
            (token_similarity(query, record.normalized_name), record)
            for record in records
        ),
        key=lambda item: (item[0], item[1].normalized_name),
    )
    best_score, best_record = ranked[-1]
    second_score = ranked[-2][0] if len(ranked) > 1 else 0.0

    if (
        best_score < FUZZY_MATCH_THRESHOLD
        or best_score - second_score < FUZZY_MATCH_MARGIN
    ):
        raise NameMatchError(
            "No se encontró una coincidencia única con suficiente similitud."
        )

    return MatchResult(best_record, "fuzzy", best_score)


class OrozcoGradeBook:
    def __init__(
        self,
        base_records: list[GradeRecord],
        language_corrections: dict[str, GradeRecord],
    ) -> None:
        self.base_records = base_records
        self.language_corrections = language_corrections

    @classmethod
    def load(
        cls,
        base_file: str | Path,
        language_file: str | Path,
    ) -> "OrozcoGradeBook":
        base_records = read_grade_records(
            base_file,
            allow_partial_notes=True,
        )
        correction_records = read_grade_records(
            language_file,
            allow_partial_notes=True,
        )
        language_corrections = {}

        for correction in correction_records:
            match = find_unique_match(correction.name, base_records)
            normalized_base_name = match.record.normalized_name
            if normalized_base_name in language_corrections:
                raise GradeFileError(
                    "Dos correcciones de Lengua apuntan al mismo estudiante."
                )
            language_corrections[normalized_base_name] = correction

        return cls(base_records, language_corrections)

    def get_notes(
        self,
        portal_student_name: object,
        subject_name: str,
    ) -> tuple[str | None, str | None, str | None, str | None]:
        match = find_unique_match(portal_student_name, self.base_records)
        base_notes = match.record.notes

        if normalize_subject(subject_name) == normalize_subject(LANGUAGE_SUBJECT):
            correction = self.language_corrections.get(match.record.normalized_name)
            if correction:
                return tuple(
                    correction_note if correction_note is not None else base_note
                    for base_note, correction_note in zip(base_notes, correction.notes)
                )

        return base_notes
